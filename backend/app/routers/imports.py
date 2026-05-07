"""Transaction import routes (CSV/Excel preview + confirm)."""

from __future__ import annotations

import csv
import hashlib
import io
import re
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, Optional

from fastapi import APIRouter, Depends, File, Form, Query, UploadFile

from ..deps import AuthContext, get_authenticated_context
from ..errors import map_value_error_to_http
from ..schemas import (
    ImportConfirmRequest,
    ImportConfirmResponse,
    ImportHistoryRecord,
    ImportPreviewResponse,
    ImportPreviewRow,
)

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional runtime dependency
    load_workbook = None  # type: ignore[assignment]

router = APIRouter(prefix="/imports", tags=["imports"])

HEADER_ALIASES = {
    "date": ["date", "transaction date", "ngay giao dich", "ngay", "ngay gd"],
    "description": ["description", "note", "noi dung", "details", "memo"],
    "amount": ["amount", "so tien", "value"],
    "type": ["type", "transaction type", "loai", "loai giao dich"],
    "debit": ["debit", "withdraw", "ghi no", "chi ra", "outflow"],
    "credit": ["credit", "deposit", "ghi co", "thu vao", "inflow"],
}

KEYWORD_GROUPS = [
    (["coffee", "highlands", "starbucks", "kfc", "restaurant", "food", "an", "cafe"], "food"),
    (["grab", "taxi", "bus", "be", "xanh sm", "transport"], "transport"),
    (["shopee", "lazada", "tiki", "mall", "shopping"], "shopping"),
    (["electric", "water", "internet", "wifi", "dien", "nuoc"], "utilities"),
    (
        ["hospital", "pharmacy", "medicine", "medical", "benh vien", "thuoc"],
        "health",
    ),
]

CATEGORY_GROUP_HINTS = {
    "food": ["food", "dining", "eat", "meal", "restaurant", "cafe", "coffee"],
    "transport": ["transport", "travel", "taxi", "bus", "ride"],
    "shopping": ["shopping", "shop", "mall", "retail"],
    "utilities": ["utilities", "utility", "electric", "water", "internet", "wifi"],
    "health": ["health", "healthcare", "medical", "hospital", "pharmacy"],
}

TYPE_HINTS_INCOME = ("income", "credit", "deposit", "thu", "salary")
TYPE_HINTS_EXPENSE = ("expense", "debit", "withdraw", "chi", "payment", "spend")


def _normalize_text(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().strip()
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_text_token(value: str) -> str:
    text = _normalize_text(value)
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _normalize_description(value: str | None) -> str:
    return re.sub(r"\s+", " ", _normalize_text(value or "")).strip()


def _decode_csv_bytes(payload: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252"):
        try:
            return payload.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Không thể đọc file CSV. Vui lòng dùng file mã hóa UTF-8.")


def _read_csv_rows(payload: bytes) -> list[dict[str, Any]]:
    text = _decode_csv_bytes(payload)
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("File CSV phải có dòng tiêu đề (header).")
    return [dict(row) for row in reader]


def _read_excel_rows(payload: bytes) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise ValueError("Tính năng nhập Excel cần cài thêm thư viện `openpyxl`.")

    workbook = load_workbook(filename=io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    output_rows: list[dict[str, Any]] = []
    for row in rows_iter:
        values = list(row)
        if not any(value is not None and str(value).strip() for value in values):
            continue
        mapped: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            mapped[header] = values[idx] if idx < len(values) else None
        output_rows.append(mapped)
    return output_rows


def _build_header_map(headers: Iterable[str]) -> dict[str, str]:
    mapping: dict[str, str] = {}
    for original in headers:
        normalized = _normalize_text_token(original)
        if normalized and normalized not in mapping:
            mapping[normalized] = original
    return mapping


def _find_column(header_map: dict[str, str], aliases: list[str]) -> Optional[str]:
    for alias in aliases:
        normalized = _normalize_text_token(alias)
        if normalized in header_map:
            return header_map[normalized]
    return None


def _parse_date(value: Any) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(value: Any) -> Optional[Decimal]:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))

    raw = str(value).strip()
    if not raw:
        return None

    is_negative = raw.startswith("(") and raw.endswith(")")
    raw = raw.strip("()")
    cleaned = re.sub(r"[^0-9,.\-]", "", raw)

    if cleaned.count(".") > 1 and "," not in cleaned:
        cleaned = cleaned.replace(".", "")
    if cleaned.count(",") > 1 and "." not in cleaned:
        cleaned = cleaned.replace(",", "")
    if "," in cleaned and "." not in cleaned:
        if re.search(r",\d{3}$", cleaned):
            cleaned = cleaned.replace(",", "")
        else:
            cleaned = cleaned.replace(",", ".")
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(",", "")
    if "." in cleaned and "," not in cleaned and re.search(r"\.\d{3}$", cleaned):
        cleaned = cleaned.replace(".", "")

    try:
        amount = Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None

    if is_negative:
        amount *= Decimal("-1")
    return amount


def _parse_type_and_amount(
    row: dict[str, Any],
    date_col: Optional[str],
    desc_col: Optional[str],
    amount_col: Optional[str],
    type_col: Optional[str],
    debit_col: Optional[str],
    credit_col: Optional[str],
) -> dict[str, Any]:
    raw_date = row.get(date_col) if date_col else None
    raw_desc = row.get(desc_col) if desc_col else None
    raw_amount = row.get(amount_col) if amount_col else None
    raw_type = row.get(type_col) if type_col else None

    parsed_date = _parse_date(raw_date)
    type_hint = _normalize_text_token(str(raw_type or ""))

    debit_amount = _parse_amount(row.get(debit_col)) if debit_col else None
    credit_amount = _parse_amount(row.get(credit_col)) if credit_col else None
    direct_amount = _parse_amount(raw_amount)

    parsed_type: Optional[str] = None
    parsed_amount: Optional[Decimal] = None

    if credit_amount and credit_amount > 0 and (not debit_amount or debit_amount == 0):
        parsed_type = "INCOME"
        parsed_amount = credit_amount
    elif debit_amount and debit_amount > 0 and (not credit_amount or credit_amount == 0):
        parsed_type = "EXPENSE"
        parsed_amount = debit_amount
    elif direct_amount is not None:
        if any(token in type_hint for token in TYPE_HINTS_INCOME):
            parsed_type = "INCOME"
            parsed_amount = abs(direct_amount)
        elif any(token in type_hint for token in TYPE_HINTS_EXPENSE):
            parsed_type = "EXPENSE"
            parsed_amount = abs(direct_amount)
        elif direct_amount < 0:
            parsed_type = "EXPENSE"
            parsed_amount = abs(direct_amount)
        elif direct_amount > 0:
            parsed_type = "INCOME"
            parsed_amount = direct_amount

    error_message = None
    if parsed_date is None:
        error_message = "Không thể phân tích ngày giao dịch."
    elif parsed_type is None or parsed_amount is None or parsed_amount <= 0:
        error_message = "Không thể suy luận loại giao dịch hoặc số tiền."

    return {
        "raw_date": str(raw_date).strip() if raw_date is not None else None,
        "raw_description": str(raw_desc).strip() if raw_desc is not None else "",
        "raw_amount": str(raw_amount).strip() if raw_amount is not None else None,
        "raw_type": str(raw_type).strip() if raw_type is not None else None,
        "parsed_date": parsed_date,
        "parsed_type": parsed_type,
        "parsed_amount": parsed_amount,
        "error_message": error_message,
    }


def _build_duplicate_hash(
    user_id: int,
    account_id: int,
    parsed_date: Optional[date],
    parsed_amount: Optional[Decimal],
    description: str | None,
    transaction_type: str | None,
) -> str:
    date_part = parsed_date.isoformat() if isinstance(parsed_date, date) else ""
    amount_part = f"{parsed_amount:.2f}" if parsed_amount is not None else ""
    normalized_desc = _normalize_description(description)
    payload = "|".join(
        [
            str(user_id),
            str(account_id),
            date_part,
            amount_part,
            normalized_desc,
            str(transaction_type or "").upper(),
        ]
    )
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _build_category_name_map(categories: list[dict[str, Any]]) -> dict[int, str]:
    return {
        int(row["CategoryID"]): str(row["CategoryName"])
        for row in categories
        if "CategoryID" in row and "CategoryName" in row
    }


def _find_category_by_group(
    categories: list[dict[str, Any]],
    group_key: str,
) -> Optional[int]:
    hints = CATEGORY_GROUP_HINTS.get(group_key, [])
    for row in categories:
        name = _normalize_text_token(str(row.get("CategoryName", "")))
        if not name:
            continue
        if any(hint in name for hint in hints):
            return int(row["CategoryID"])
    return None


def _load_category_rules(
    ctx: AuthContext,
    target_user_id: int,
) -> list[dict[str, Any]]:
    query = """
        SELECT RuleID, UserID, Keyword, CategoryID, Priority
        FROM TransactionCategoryRules
        WHERE IsActive = 1
          AND (UserID IS NULL OR UserID = %s)
        ORDER BY
            CASE WHEN UserID = %s THEN 0 ELSE 1 END,
            Priority,
            RuleID
    """
    cursor = ctx.service.connection.cursor(dictionary=True)
    try:
        cursor.execute(query, (target_user_id, target_user_id))
        return cursor.fetchall()
    except Exception:
        return []
    finally:
        cursor.close()


def _suggest_category_id(
    description: str,
    categories: list[dict[str, Any]],
    category_name_map: dict[int, str],
    rules: list[dict[str, Any]],
) -> Optional[int]:
    normalized = _normalize_description(description)
    if not normalized:
        return None

    for rule in rules:
        keyword = _normalize_description(str(rule.get("Keyword") or ""))
        category_id = rule.get("CategoryID")
        if not keyword or category_id is None:
            continue
        category_id = int(category_id)
        if category_id not in category_name_map:
            continue
        if keyword in normalized:
            return category_id

    for keywords, group_key in KEYWORD_GROUPS:
        if any(keyword in normalized for keyword in keywords):
            return _find_category_by_group(categories, group_key)

    return None


def _resolve_target_user_id(ctx: AuthContext, requested_user_id: Optional[int]) -> int:
    auth_user_id = int(ctx.user["UserID"])
    role = str(ctx.user.get("UserRole", "USER")).upper()

    if requested_user_id is None:
        return auth_user_id

    requested = int(requested_user_id)
    if requested == auth_user_id:
        return requested
    if role != "ADMIN":
        raise ValueError("Bạn chỉ có thể nhập dữ liệu cho tài khoản của chính mình.")
    if not ctx.service.get_user_by_id(requested):
        raise ValueError(f"Không tồn tại người dùng có mã {requested}.")
    return requested


def _load_raw_rows(file_name: str, payload: bytes) -> tuple[list[dict[str, Any]], str]:
    lower_name = (file_name or "").lower()
    if lower_name.endswith(".csv"):
        return _read_csv_rows(payload), "CSV"
    if lower_name.endswith(".xlsx") or lower_name.endswith(".xlsm") or lower_name.endswith(".xls"):
        return _read_excel_rows(payload), "EXCEL"
    raise ValueError("Định dạng file chưa được hỗ trợ. Vui lòng tải lên CSV hoặc Excel.")


@router.post("/transactions/preview", response_model=ImportPreviewResponse)
async def preview_transactions_import(
    file: UploadFile = File(...),
    account_id: int = Form(...),
    user_id: Optional[int] = Form(default=None),
    ctx: AuthContext = Depends(get_authenticated_context),
) -> ImportPreviewResponse:
    try:
        target_user_id = _resolve_target_user_id(ctx, user_id)
        if not ctx.service.account_belongs_to_user(target_user_id, account_id):
            raise ValueError(
                f"Tài khoản {account_id} không thuộc người dùng {target_user_id}."
            )

        payload = await file.read()
        raw_rows, file_type = _load_raw_rows(file.filename or "", payload)
        if not raw_rows:
            raise ValueError("File nhập không có dòng dữ liệu.")

        header_map = _build_header_map(raw_rows[0].keys())
        date_col = _find_column(header_map, HEADER_ALIASES["date"])
        desc_col = _find_column(header_map, HEADER_ALIASES["description"])
        amount_col = _find_column(header_map, HEADER_ALIASES["amount"])
        type_col = _find_column(header_map, HEADER_ALIASES["type"])
        debit_col = _find_column(header_map, HEADER_ALIASES["debit"])
        credit_col = _find_column(header_map, HEADER_ALIASES["credit"])

        if not date_col:
            raise ValueError("Không tìm thấy cột ngày giao dịch trong file.")
        if not desc_col:
            raise ValueError("Không tìm thấy cột mô tả/nội dung trong file.")
        if not amount_col and not (debit_col or credit_col):
            raise ValueError("Không tìm thấy cột số tiền hoặc ghi nợ/ghi có trong file.")

        categories = ctx.service.list_categories()
        category_name_map = _build_category_name_map(categories)
        rules = _load_category_rules(ctx, target_user_id)

        existing_hashes: set[str] = set()
        existing_rows = ctx.service.list_transactions(
            user_id=target_user_id,
            account_id=account_id,
        )
        for row in existing_rows:
            existing_hashes.add(
                _build_duplicate_hash(
                    user_id=target_user_id,
                    account_id=account_id,
                    parsed_date=row.get("TransactionDate").date()
                    if isinstance(row.get("TransactionDate"), datetime)
                    else None,
                    parsed_amount=Decimal(str(row.get("Amount"))) if row.get("Amount") is not None else None,
                    description=row.get("Description"),
                    transaction_type=row.get("TransactionType"),
                )
            )

        cursor = ctx.service.connection.cursor(dictionary=True)
        staging_hashes: set[str] = set()
        preview_rows: list[ImportPreviewRow] = []
        try:
            cursor.execute(
                """
                INSERT INTO TransactionImportBatches
                    (UserID, AccountID, FileName, FileType, Status, TotalRows)
                VALUES (%s, %s, %s, %s, 'PREVIEWED', %s)
                """,
                (target_user_id, account_id, file.filename, file_type, len(raw_rows)),
            )
            batch_id = int(cursor.lastrowid)

            for index, raw_row in enumerate(raw_rows, start=1):
                parsed = _parse_type_and_amount(
                    row=raw_row,
                    date_col=date_col,
                    desc_col=desc_col,
                    amount_col=amount_col,
                    type_col=type_col,
                    debit_col=debit_col,
                    credit_col=credit_col,
                )
                duplicate_hash = _build_duplicate_hash(
                    user_id=target_user_id,
                    account_id=account_id,
                    parsed_date=parsed["parsed_date"],
                    parsed_amount=parsed["parsed_amount"],
                    description=parsed["raw_description"],
                    transaction_type=parsed["parsed_type"],
                )
                is_duplicate = int(
                    duplicate_hash in existing_hashes or duplicate_hash in staging_hashes
                )
                if duplicate_hash:
                    staging_hashes.add(duplicate_hash)

                suggested_category_id = None
                if parsed["parsed_type"] == "EXPENSE":
                    suggested_category_id = _suggest_category_id(
                        description=parsed["raw_description"] or "",
                        categories=categories,
                        category_name_map=category_name_map,
                        rules=rules,
                    )

                action = "SKIP" if is_duplicate else "IMPORT"
                error_message = parsed["error_message"]
                if is_duplicate and not error_message:
                    error_message = "Có khả năng trùng giao dịch."
                if parsed["parsed_type"] is None or parsed["parsed_amount"] is None:
                    action = "SKIP"

                cursor.execute(
                    """
                    INSERT INTO TransactionImportRows (
                        BatchID,
                        RowNumber,
                        RawDate,
                        RawDescription,
                        RawAmount,
                        RawType,
                        ParsedDate,
                        ParsedAmount,
                        ParsedType,
                        SuggestedCategoryID,
                        IsDuplicate,
                        DuplicateHash,
                        ActionType,
                        RowStatus,
                        ErrorMessage
                    )
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'PREVIEW', %s)
                    """,
                    (
                        batch_id,
                        index,
                        parsed["raw_date"],
                        parsed["raw_description"],
                        parsed["raw_amount"],
                        parsed["raw_type"],
                        parsed["parsed_date"],
                        float(parsed["parsed_amount"]) if parsed["parsed_amount"] is not None else None,
                        parsed["parsed_type"],
                        suggested_category_id,
                        is_duplicate,
                        duplicate_hash,
                        action,
                        error_message,
                    ),
                )
                row_id = int(cursor.lastrowid)
                preview_rows.append(
                    ImportPreviewRow(
                        row_id=row_id,
                        row_number=index,
                        raw_date=parsed["raw_date"],
                        raw_description=parsed["raw_description"],
                        raw_amount=parsed["raw_amount"],
                        raw_type=parsed["raw_type"],
                        parsed_date=parsed["parsed_date"],
                        parsed_amount=float(parsed["parsed_amount"])
                        if parsed["parsed_amount"] is not None
                        else None,
                        parsed_type=parsed["parsed_type"],
                        suggested_category_id=suggested_category_id,
                        suggested_category_name=category_name_map.get(suggested_category_id)
                        if suggested_category_id is not None
                        else None,
                        is_duplicate=is_duplicate,
                        action=action,
                        error_message=error_message,
                    )
                )

            ctx.service.connection.commit()
        finally:
            cursor.close()

        return ImportPreviewResponse(
            batch_id=batch_id,
            status="PREVIEWED",
            total_rows=len(preview_rows),
            rows=preview_rows,
        )
    except ValueError as err:
        raise map_value_error_to_http(err, default_status=400) from err


@router.post("/transactions/confirm", response_model=ImportConfirmResponse)
def confirm_transactions_import(
    payload: ImportConfirmRequest,
    ctx: AuthContext = Depends(get_authenticated_context),
) -> ImportConfirmResponse:
    try:
        target_user_id = _resolve_target_user_id(ctx, payload.user_id)
        categories = ctx.service.list_categories()
        valid_category_ids = {int(row["CategoryID"]) for row in categories}

        cursor = ctx.service.connection.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT BatchID, UserID, AccountID, Status
                FROM TransactionImportBatches
                WHERE BatchID = %s
                """,
                (payload.batch_id,),
            )
            batch = cursor.fetchone()
            if not batch:
                raise ValueError(f"Không tồn tại lô nhập có mã {payload.batch_id}.")
            if int(batch["UserID"]) != target_user_id:
                raise ValueError("Lô nhập không thuộc phạm vi người dùng đang chọn.")

            for row_update in payload.rows:
                action = (row_update.action or "IMPORT").strip().upper()
                if action not in {"IMPORT", "SKIP"}:
                    action = "SKIP"
                if (
                    row_update.final_category_id is not None
                    and int(row_update.final_category_id) not in valid_category_ids
                ):
                    raise ValueError(
                        f"Không tồn tại danh mục có mã {row_update.final_category_id}."
                    )
                cursor.execute(
                    """
                    UPDATE TransactionImportRows
                    SET ActionType = %s,
                        FinalCategoryID = %s
                    WHERE BatchID = %s
                      AND RowID = %s
                    """,
                    (
                        action,
                        row_update.final_category_id,
                        payload.batch_id,
                        row_update.row_id,
                    ),
                )

            cursor.execute(
                """
                SELECT
                    RowID,
                    RawDescription,
                    ParsedDate,
                    ParsedAmount,
                    ParsedType,
                    SuggestedCategoryID,
                    FinalCategoryID,
                    IsDuplicate,
                    ActionType
                FROM TransactionImportRows
                WHERE BatchID = %s
                ORDER BY RowNumber
                """,
                (payload.batch_id,),
            )
            import_rows = cursor.fetchall()

            imported_rows = 0
            skipped_rows = 0
            failed_rows = 0
            account_id = int(batch["AccountID"])

            for row in import_rows:
                row_id = int(row["RowID"])
                action = str(row.get("ActionType") or "IMPORT").upper()
                parsed_type = str(row.get("ParsedType") or "").upper()
                parsed_date = row.get("ParsedDate")
                parsed_amount = row.get("ParsedAmount")
                description = (row.get("RawDescription") or "").strip()
                error_message = None

                if int(row.get("IsDuplicate") or 0) == 1:
                    action = "SKIP"
                    error_message = "Đã bỏ qua dòng bị trùng."

                if action != "IMPORT":
                    skipped_rows += 1
                    cursor.execute(
                        """
                        UPDATE TransactionImportRows
                        SET RowStatus = 'SKIPPED',
                            ErrorMessage = COALESCE(%s, ErrorMessage)
                        WHERE RowID = %s
                        """,
                        (error_message, row_id),
                    )
                    continue

                if parsed_date is None or parsed_amount is None or parsed_type not in {"INCOME", "EXPENSE"}:
                    failed_rows += 1
                    cursor.execute(
                        """
                        UPDATE TransactionImportRows
                        SET RowStatus = 'FAILED',
                            ErrorMessage = 'Thiếu dữ liệu ngày, loại giao dịch hoặc số tiền sau khi phân tích.'
                        WHERE RowID = %s
                        """,
                        (row_id,),
                    )
                    continue

                try:
                    amount = float(parsed_amount)
                    if parsed_type == "INCOME":
                        ctx.service.create_income(
                            user_id=target_user_id,
                            account_id=account_id,
                            amount=amount,
                            description=description,
                            transaction_date=parsed_date,
                        )
                    else:
                        category_id = row.get("FinalCategoryID") or row.get("SuggestedCategoryID")
                        if category_id is None:
                            raise ValueError("Thiếu danh mục cho dòng chi tiêu.")
                        ctx.service.create_expense(
                            user_id=target_user_id,
                            account_id=account_id,
                            category_id=int(category_id),
                            amount=amount,
                            description=description,
                            transaction_date=parsed_date,
                        )

                    imported_rows += 1
                    cursor.execute(
                        """
                        UPDATE TransactionImportRows
                        SET RowStatus = 'IMPORTED',
                            ErrorMessage = NULL
                        WHERE RowID = %s
                        """,
                        (row_id,),
                    )
                except Exception as err:
                    failed_rows += 1
                    cursor.execute(
                        """
                        UPDATE TransactionImportRows
                        SET RowStatus = 'FAILED',
                            ErrorMessage = %s
                        WHERE RowID = %s
                        """,
                        (str(err)[:255], row_id),
                    )

            batch_status = "CONFIRMED"
            if failed_rows > 0 and imported_rows > 0:
                batch_status = "PARTIAL"
            elif failed_rows > 0 and imported_rows == 0:
                batch_status = "FAILED"

            cursor.execute(
                """
                UPDATE TransactionImportBatches
                SET Status = %s,
                    ImportedRows = %s,
                    SkippedRows = %s,
                    FailedRows = %s,
                    ConfirmedAt = CURRENT_TIMESTAMP
                WHERE BatchID = %s
                """,
                (batch_status, imported_rows, skipped_rows, failed_rows, payload.batch_id),
            )
            ctx.service.connection.commit()
        finally:
            cursor.close()

        return ImportConfirmResponse(
            message="Nhập giao dịch hoàn tất.",
            batch_id=payload.batch_id,
            imported_rows=imported_rows,
            skipped_rows=skipped_rows,
            failed_rows=failed_rows,
        )
    except ValueError as err:
        raise map_value_error_to_http(err, default_status=400) from err


@router.get("/transactions/history", response_model=list[ImportHistoryRecord])
def get_import_history(
    user_id: Optional[int] = Query(default=None),
    ctx: AuthContext = Depends(get_authenticated_context),
) -> list[ImportHistoryRecord]:
    try:
        role = str(ctx.user.get("UserRole", "USER")).upper()
        if role != "ADMIN":
            target_user_id = int(ctx.user["UserID"])
        elif user_id is None:
            target_user_id = None
        else:
            target_user_id = _resolve_target_user_id(ctx, user_id)

        cursor = ctx.service.connection.cursor(dictionary=True)
        try:
            cursor.execute(
                """
                SELECT
                    b.BatchID AS batch_id,
                    b.UserID AS user_id,
                    b.AccountID AS account_id,
                    banks.BankName AS bank_name,
                    b.FileName AS file_name,
                    b.FileType AS file_type,
                    b.Status AS status,
                    b.TotalRows AS total_rows,
                    b.ImportedRows AS imported_rows,
                    b.SkippedRows AS skipped_rows,
                    b.FailedRows AS failed_rows,
                    b.CreatedAt AS created_at,
                    b.ConfirmedAt AS confirmed_at
                FROM TransactionImportBatches b
                JOIN BankAccounts accounts
                  ON accounts.AccountID = b.AccountID
                 AND accounts.UserID = b.UserID
                JOIN Banks banks
                  ON banks.BankID = accounts.BankID
                WHERE (%s IS NULL OR b.UserID = %s)
                ORDER BY b.BatchID DESC
                LIMIT 50
                """,
                (target_user_id, target_user_id),
            )
            rows = cursor.fetchall()
        finally:
            cursor.close()

        return [ImportHistoryRecord(**row) for row in rows]
    except ValueError as err:
        raise map_value_error_to_http(err, default_status=400) from err
